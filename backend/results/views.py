from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from .serializers import StudentProfileSerializer
from .models import ResultUpload, ResultEntry
from attendance.models import StaffProfile, StudentProfile
from openpyxl import load_workbook
from django.conf import settings
from attendance.models import Attendance
# from attendance.models import SubjectAssignment
# from django.db.models import Count, Q
from django.db.models import Sum
from django.contrib.auth import get_user_model

User = get_user_model()




# ==============================================================
# RESULT CONTEXT AND UPLOAD PREVIEW
# ==============================================================

class ResultUploadCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):

        staff = StaffProfile.objects.filter(user=request.user).first()
        if not staff:
            return Response({"detail": "Unauthorized"}, status=403)

        file = request.FILES.get("file")
        batch_id = request.data.get("batch_id")

        if not file or not batch_id:
            return Response(
                {"detail": "File and batch_id are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        ResultUpload.objects.filter(batch_id=batch_id).delete()

        upload = ResultUpload.objects.create(
            file=file,
            uploaded_by=request.user,
            batch_id=batch_id
        )

        file_path = upload.file.path
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        preview_data = []
        unique_students = set()
        total_percentage_list = []
        passed = 0
        failed = 0

        def parse_marks(value):
            if not value or value == "NA":
                return None, None

            if isinstance(value, str) and "/" in value:
                obtained, total = value.split("/")
                return float(obtained.strip()), float(total.strip())

            return None, None

        for row in sheet.iter_rows(min_row=2, values_only=True):

            roll_number = str(row[1]).strip() if row[1] else None  # College ID
            subject_name = str(row[3]).strip() if row[3] else "-"
            internal_raw = row[4]
            external_raw = row[5]

            if not roll_number:
                continue

            unique_students.add(roll_number)

            internal_obt, internal_total = parse_marks(internal_raw)
            external_obt, external_total = parse_marks(external_raw)

            # ===============================
            # CALCULATE TOTAL FROM INTERNAL + EXTERNAL
            # ===============================

            total_obt = None
            total_total = None

            if internal_obt is not None and external_obt is not None:
                total_obt = internal_obt + external_obt
                total_total = internal_total + external_total

            # ===============================
            # FORMAT DISPLAY AS 00/00
            # ===============================

            internal_display = "-" if internal_obt is None else f"{int(internal_obt)}/{int(internal_total)}"
            external_display = "-" if external_obt is None else f"{int(external_obt)}/{int(external_total)}"
            total_display = "-" if total_obt is None else f"{int(total_obt)}/{int(total_total)}"

            grade = "F"
            status_label = "Fail"

            # If any side missing → FAIL
            if internal_obt is None or external_obt is None:
                grade = "F"
                status_label = "Fail"
                failed += 1

            elif total_obt is not None and total_total:
                percentage = (total_obt / total_total) * 100
                total_percentage_list.append(percentage)

                if percentage >= 95:
                    grade = "A+"
                elif percentage >= 85:
                    grade = "A"
                elif percentage >= 75:
                    grade = "B"
                elif percentage >= 65:
                    grade = "C"
                elif percentage >= 55:
                    grade = "P"
                else:
                    grade = "F"

                if grade != "F":
                    passed += 1
                    status_label = "Pass"
                else:
                    failed += 1

            preview_data.append({
                "rollNumber": roll_number,
                "subjectName": subject_name,  # Showing subject name as requested
                "internalMarks": internal_display,
                "externalMarks": external_display,
                "totalMarks": total_display,
                "maxMarks": total_total if total_total else 0,
                "grade": grade,
                "status": status_label,
            })

        total_students = len(unique_students)

        average_marks = (
            round(sum(total_percentage_list) / len(total_percentage_list))
            if total_percentage_list else 0
        )

        summary = {
            "totalStudents": total_students,
            "passed": passed,
            "failed": failed,
            "averageMarks": average_marks,
        }

        return Response({
            "upload_id": upload.upload_id,
            "preview": preview_data,
            "summary": summary
        }, status=status.HTTP_201_CREATED)
    


# ==============================================================
# DATABASE PUBLISHING VIEW
# ==============================================================
class ResultPublishView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, upload_id):

        upload = ResultUpload.objects.filter(upload_id=upload_id).first()

        if not upload:
            return Response({"detail": "Upload not found"}, status=404)

        file_path = upload.file.path
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        def parse_marks(value):
            if not value or value == "NA":
                return None, None
            if isinstance(value, str) and "/" in value:
                obtained, total = value.split("/")
                return float(obtained.strip()), float(total.strip())
            return None, None

        for row in sheet.iter_rows(min_row=2, values_only=True):

            roll_number = str(row[1]).strip() if row[1] else None
            subject_name = str(row[3]).strip() if row[3] else "-"
            internal_raw = row[4]
            external_raw = row[5]
            credits = row[6] if row[6] else 0
            sgpa = row[7] if row[7] else None

            if not roll_number:
                continue

            internal_obt, internal_total = parse_marks(internal_raw)
            external_obt, external_total = parse_marks(external_raw)

            total_display = "-"
            grade = "F"

            if internal_obt is not None and external_obt is not None:
                total_obt = internal_obt + external_obt
                total_total = internal_total + external_total

                total_display = f"{int(total_obt)}/{int(total_total)}"
                percentage = (total_obt / total_total) * 100

                if percentage >= 95:
                    grade = "A+"
                elif percentage >= 85:
                    grade = "A"
                elif percentage >= 75:
                    grade = "B"
                elif percentage >= 65:
                    grade = "C"
                elif percentage >= 55:
                    grade = "P"
                else:
                    grade = "F"
            try:
                student_user = settings.AUTH_USER_MODEL.objects.get(username=roll_number)
            except Exception:
                from django.contrib.auth import get_user_model
                User = get_user_model()
                try:
                    student_user = User.objects.get(username=roll_number)
                except User.DoesNotExist:
                    continue  # Skip if student not found

            ResultEntry.objects.create(
                upload=upload,
                batch_id=upload.batch_id,
                student=student_user,   # ✅ correct foreign key
                subject_code=str(row[2]).strip() if row[2] else "-",
                subject_name=subject_name,
                internal_marks=internal_raw if internal_raw else "-",
                external_marks=external_raw if external_raw else "-",
                total_marks=total_display,
                grade=grade,
                subject_credits=float(row[6]) if row[6] else 0,
                sgpa=float(row[7]) if row[7] else 0,
            )

        return Response({"detail": "Results published successfully"})



# ==============================================================
# BATCH PERFORMANCE VIEW FOR ANALYTICS AND CHARTS
# ==============================================================
class BatchPerformanceView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):

        batches = ResultUpload.objects.values_list("batch_id", flat=True).distinct()

        performance_data = []

        for batch in batches:

            # -----------------------------
            # ATTENDANCE
            # -----------------------------
            attendance_records = Attendance.objects.filter(
                teaching_plan__assignment__batch_id=batch
            )

            total_attendance = attendance_records.count()
            total_present = attendance_records.filter(is_present=True).count()

            avg_attendance = 0
            if total_attendance > 0:
                avg_attendance = round((total_present / total_attendance) * 100)

            # -----------------------------
            # RESULTS
            # -----------------------------
            results = ResultEntry.objects.filter(batch_id=batch)

            total_obtained = 0
            total_outof = 0

            for result in results:
                if result.total_marks and "/" in result.total_marks:
                    obtained, outof = result.total_marks.split("/")
                    total_obtained += float(obtained)
                    total_outof += float(outof)

            avg_score = 0
            if total_outof > 0:
                avg_score = round((total_obtained / total_outof) * 100)

            performance_data.append({
                "batch": batch,
                "averageAttendance": avg_attendance,
                "averageScore": avg_score,
            })

        return Response(performance_data)
    






# ==============================================================
# STUDENT RESULT PAGE VIEW FUNCTIONS
# ==============================================================


# ==============================================================
# RESULT SUMMARY CARDS FOUR (FOR STUDENT DASHBOARD)
# ==============================================================
class StudentResultSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):

        user = request.user

        # -----------------------------------
        # 1️⃣ Get Student Profile (Semester)
        # -----------------------------------
        student_profile = StudentProfile.objects.filter(user=user).first()

        if not student_profile:
            return Response({"detail": "Student profile not found"}, status=404)

        current_sem = int(student_profile.semester)

        # -----------------------------------
        # 2️⃣ Get Results of Current Semester
        # -----------------------------------
        results = ResultEntry.objects.filter(student=user).order_by("-id")

        if not results.exists():
            return Response({
                "latestSemester": current_sem,
                "currentSgpa": 0,
                "totalBacklogs": 0,
                "creditsEarned": 0,
                "totalCredits": 0,
            })

        # -----------------------------------
        # 3️⃣ SGPA (take latest SGPA entry)
        # -----------------------------------
        latest_result = results.first()
        current_sgpa = latest_result.sgpa if latest_result.sgpa else 0

        # -----------------------------------
        # 4️⃣ Backlogs (Grade = F)
        # -----------------------------------
        total_backlogs = results.filter(grade="F").count()

        # -----------------------------------
        # 5️⃣ Credits
        # -----------------------------------
        total_credits = results.aggregate(
            total=Sum("subject_credits")
        )["total"] or 0

        earned_credits = results.exclude(grade="F").aggregate(
            total=Sum("subject_credits")
        )["total"] or 0

        return Response({
            "latestSemester": current_sem,
            "currentSgpa": round(current_sgpa, 2),
            "totalBacklogs": total_backlogs,
            "creditsEarned": earned_credits,
            "totalCredits": total_credits,
        })
    


# ==============================================================
# SUBJECT WISE DISCRIPTION TABLE (FOR STUDENT DASHBOARD)
# ==============================================================

class StudentSemesterResultView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):

        user = request.user

        results = ResultEntry.objects.filter(student=user)

        if not results.exists():
            return Response({
    "subjects": [],
    "isPublished": False
})

        subjects = []

        for r in results:

            def parse_marks(value):
                if value and "/" in value:
                    obtained, max_val = value.split("/")
                    return int(obtained), int(max_val)
                return 0, 0

            internal_obt, internal_max = parse_marks(r.internal_marks)
            external_obt, external_max = parse_marks(r.external_marks)

            total_obt = 0
            total_max = 0

            if r.total_marks and "/" in r.total_marks:
                total_obt, total_max = map(int, r.total_marks.split("/"))

            subjects.append({
                    # 🔥 ADDED REQUIRED FIELDS
                    "code": r.subject_code,
                    "name": r.subject_name,
                    "credits": float(r.subject_credits or 0),

                    # Marks
                    "internalMarks": internal_obt,
                    "internalMax": internal_max,
                    "externalMarks": external_obt,
                    "externalMax": external_max,
                    "totalMarks": total_obt,
                    "maxMarks": total_max,

                    # Result
                    "grade": r.grade,
                    "status": "Pass" if r.grade != "F" else "Fail",
                })

        total_credits = results.aggregate(
            total=Sum("subject_credits")
        )["total"] or 0

        earned_credits = results.exclude(
            grade="F"
        ).aggregate(
            total=Sum("subject_credits")
        )["total"] or 0

        return Response({
            "examType": "Regular",
            "sgpa": results.first().sgpa or 0,
            "totalCredits": total_credits,
            "earnedCredits": earned_credits,
            "subjects": subjects,
            "isPublished": True,
        })



# ==============================================================
# Student Info (FOR MARKSHEET PDF)
# ==============================================================
class StudentProfileDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        profile = StudentProfile.objects.get(user=request.user)
        serializer = StudentProfileSerializer(profile)
        return Response(serializer.data)



# ==============================================================
# LOAD BACKLOGS (FAILED SUBJECTS) FOR STUDENT
# ==============================================================
class StudentBacklogView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):

        # ✅ Filter only failed subjects for logged-in user
        failed_entries = ResultEntry.objects.filter(
            student=request.user,
            grade="F"
        )

        backlogs = []

        for entry in failed_entries:
            backlogs.append({
                "subjectCode": entry.subject_code,
                "subjectName": entry.subject_name,
                "status": "Pending"
            })

        return Response(backlogs)